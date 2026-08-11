# -*- coding: utf-8 -*-
"""
📡 Aplikasi Peramalan Jumlah Pemasangan WiFi IndiBiz
Metode: Holt's Double Exponential Smoothing ATAU Holt-Winters (dengan musiman)
Deploy dengan Streamlit.

Alur: Pengumpulan Data -> Preprocessing (Data Cleaning) -> Penentuan Periode ->
      Pembagian Data (Train-Test) -> Pemilihan Metode Peramalan (ADF + pilihan user) ->
      Membangun Model -> Optimasi Parameter -> Evaluasi Model (MAE, RMSE, MAPE) ->
      Pemilihan Model Terbaik -> Peramalan ke Depan
"""

import tempfile

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from statsmodels.tsa.stattools import adfuller
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import streamlit as st

# ──────────────────────────────────────────────────────────────────────────
# 1. FUNGSI-FUNGSI INTI (metrik umum)
# ──────────────────────────────────────────────────────────────────────────

def hitung_metrik(aktual, ramalan):
    aktual = np.array(aktual, dtype=float)
    ramalan = np.array(ramalan, dtype=float)
    error = aktual - ramalan

    mae = np.mean(np.abs(error))
    rmse = np.sqrt(np.mean(error ** 2))

    mask = aktual != 0
    mape = np.mean(np.abs(error[mask] / aktual[mask])) * 100 if mask.sum() > 0 else np.nan

    return mae, rmse, mape


def kategori_mape(mape):
    if mape <= 10:
        return "Sangat Baik"
    elif mape <= 20:
        return "Baik"
    elif mape <= 50:
        return "Cukup Baik"
    else:
        return "Buruk"


# ──────────────────────────────────────────────────────────────────────────
# 2. METODE 1 — Holt's Double Exponential Smoothing (tanpa musiman)
# ──────────────────────────────────────────────────────────────────────────

def holt_des_fit(data, alpha, gamma):
    """Fitting Holt's DES pada sebuah deret data."""
    X = list(data)
    n = len(X)
    St = [0.0] * n
    bt = [0.0] * n
    Ft = [None] * n

    St[0] = X[0]
    bt[0] = ((X[1] - X[0]) + (X[3] - X[2])) / 2

    for t in range(1, n):
        St[t] = alpha * X[t] + (1 - alpha) * (St[t - 1] + bt[t - 1])
        bt[t] = gamma * (St[t] - St[t - 1]) + (1 - gamma) * bt[t - 1]
        Ft[t] = St[t - 1] + bt[t - 1]

    return St, bt, Ft


def forecast_m_langkah_des(St_akhir, bt_akhir, m_langkah):
    return [St_akhir + bt_akhir * m for m in range(1, m_langkah + 1)]


def run_holt_des(bulanan, train, test, n_forecast):
    """Optimasi 81 kombinasi α,γ (dievaluasi pada data test), lalu refit ke seluruh data."""
    alpha_vals = [round(x * 0.1, 1) for x in range(1, 10)]
    gamma_vals = [round(x * 0.1, 1) for x in range(1, 10)]

    hasil_metrik = {}
    tabel_mape = pd.DataFrame(index=alpha_vals, columns=gamma_vals, dtype=float)

    for a in alpha_vals:
        for g in gamma_vals:
            St_tr, bt_tr, _ = holt_des_fit(train.values, a, g)
            ramalan_test = forecast_m_langkah_des(St_tr[-1], bt_tr[-1], len(test))
            mae, rmse, mape = hitung_metrik(test.values, ramalan_test)
            hasil_metrik[(a, g)] = (mae, rmse, mape)
            tabel_mape.loc[a, g] = round(mape, 2)

    best_key = min(hasil_metrik, key=lambda k: hasil_metrik[k][2])
    best_alpha, best_gamma = best_key
    best_mae, best_rmse, best_mape = hasil_metrik[best_key]

    St_tr, bt_tr, _ = holt_des_fit(train.values, best_alpha, best_gamma)
    ramalan_test = forecast_m_langkah_des(St_tr[-1], bt_tr[-1], len(test))
    df_eval = pd.DataFrame({
        "Periode": [t.strftime("%B %Y") for t in test.index],
        "Aktual": test.values,
        "Ramalan": [round(v, 1) for v in ramalan_test],
    })
    df_eval["Error"] = (df_eval["Aktual"] - df_eval["Ramalan"]).round(1)
    df_eval["APE (%)"] = (df_eval["Error"].abs() / df_eval["Aktual"].replace(0, np.nan) * 100).round(2)

    St, bt, Ft = holt_des_fit(bulanan.values, best_alpha, best_gamma)
    X = bulanan.values.tolist()
    periode = bulanan.index.tolist()
    n = len(X)

    rows = []
    for t in range(n):
        ft = round(Ft[t], 2) if Ft[t] is not None else None
        pet = None
        if t > 0 and X[t] != 0 and Ft[t] is not None:
            pet = round(abs((X[t] - Ft[t]) / X[t]) * 100, 2)
        rows.append({"t": t + 1, "Periode": periode[t].strftime("%b %Y"), "Xt (Aktual)": X[t],
                      "Ft (Fitted)": ft, "PEt (%)": pet})
    df_tabel = pd.DataFrame(rows)

    St_last, bt_last = St[-1], bt[-1]
    tgl_forecast = pd.date_range(start=bulanan.index[-1] + pd.DateOffset(months=1),
                                  periods=n_forecast, freq="ME")
    hasil_forecast = [max(0, round(St_last + bt_last * m)) for m in range(1, n_forecast + 1)]
    df_forecast = pd.DataFrame({"m": range(1, n_forecast + 1),
                                 "Periode": [t.strftime("%B %Y") for t in tgl_forecast],
                                 "Prediksi": hasil_forecast})

    Ft_plot = [None] + [Ft[t] for t in range(1, n)]

    return {
        "method_name": "Holt's Double Exponential Smoothing",
        "method_short": "Holt's DES",
        "mae": best_mae, "rmse": best_rmse, "mape": best_mape,
        "params_info": {"α (alpha)": best_alpha, "γ (gamma)": best_gamma},
        "df_eval": df_eval, "df_tabel": df_tabel, "df_forecast": df_forecast,
        "tabel_mape": tabel_mape, "best_alpha": best_alpha, "best_gamma": best_gamma,
        "X": X, "Ft_plot": Ft_plot, "n": n,
        "tgl_forecast": tgl_forecast, "hasil_forecast": hasil_forecast,
    }


# ──────────────────────────────────────────────────────────────────────────
# 3. METODE 2 — Holt-Winters (dengan komponen musiman)
# ──────────────────────────────────────────────────────────────────────────

def run_holt_winters(bulanan, train, test, n_forecast, seasonal_periods=12):
    """Coba 4 kombinasi (musiman aditif/multiplikatif x damped trend ya/tidak),
    dievaluasi pada data test, lalu refit kombinasi terbaik ke seluruh data."""
    if len(train) < 2 * seasonal_periods:
        return None

    kandidat = []
    for seasonal_type in ["add", "mul"]:
        for damped in [False, True]:
            try:
                model = ExponentialSmoothing(
                    train.values, trend="add", damped_trend=damped,
                    seasonal=seasonal_type, seasonal_periods=seasonal_periods,
                    initialization_method="estimated",
                )
                fit_tr = model.fit(optimized=True)
                fc_test = fit_tr.forecast(len(test))
                mae, rmse, mape = hitung_metrik(test.values, fc_test)
                kandidat.append({"seasonal": seasonal_type, "damped": damped,
                                  "mae": mae, "rmse": rmse, "mape": mape, "fit_tr": fit_tr})
            except Exception:
                continue

    if not kandidat:
        return None

    best = min(kandidat, key=lambda d: d["mape"])

    tabel_kandidat = pd.DataFrame([
        {"Musiman": "Aditif" if k["seasonal"] == "add" else "Multiplikatif",
         "Damped Trend": "Ya" if k["damped"] else "Tidak",
         "MAE": round(k["mae"], 2), "RMSE": round(k["rmse"], 2), "MAPE (%)": round(k["mape"], 2)}
        for k in kandidat
    ]).sort_values("MAPE (%)").reset_index(drop=True)

    fc_test_best = best["fit_tr"].forecast(len(test))
    df_eval = pd.DataFrame({
        "Periode": [t.strftime("%B %Y") for t in test.index],
        "Aktual": test.values,
        "Ramalan": [round(v, 1) for v in fc_test_best],
    })
    df_eval["Error"] = (df_eval["Aktual"] - df_eval["Ramalan"]).round(1)
    df_eval["APE (%)"] = (df_eval["Error"].abs() / df_eval["Aktual"].replace(0, np.nan) * 100).round(2)

    final_model = ExponentialSmoothing(
        bulanan.values, trend="add", damped_trend=best["damped"],
        seasonal=best["seasonal"], seasonal_periods=seasonal_periods,
        initialization_method="estimated",
    )
    final_fit = final_model.fit(optimized=True)

    X = bulanan.values.tolist()
    periode = bulanan.index.tolist()
    n = len(X)
    Ft_full = final_fit.fittedvalues

    rows = []
    for t in range(n):
        ft = round(Ft_full[t], 2)
        pet = round(abs((X[t] - Ft_full[t]) / X[t]) * 100, 2) if X[t] != 0 else None
        rows.append({"t": t + 1, "Periode": periode[t].strftime("%b %Y"), "Xt (Aktual)": X[t],
                      "Ft (Fitted)": ft, "PEt (%)": pet})
    df_tabel = pd.DataFrame(rows)

    tgl_forecast = pd.date_range(start=bulanan.index[-1] + pd.DateOffset(months=1),
                                  periods=n_forecast, freq="ME")
    forecast_vals = final_fit.forecast(n_forecast)
    hasil_forecast = [max(0, round(v)) for v in forecast_vals]
    df_forecast = pd.DataFrame({"m": range(1, n_forecast + 1),
                                 "Periode": [t.strftime("%B %Y") for t in tgl_forecast],
                                 "Prediksi": hasil_forecast})

    p = final_fit.params
    params_info = {
        "Musiman": "Aditif" if best["seasonal"] == "add" else "Multiplikatif",
        "Damped Trend": "Ya" if best["damped"] else "Tidak",
        "α (level)": round(p["smoothing_level"], 3),
        "β (trend)": round(p["smoothing_trend"], 3),
        "γ (musiman)": round(p["smoothing_seasonal"], 3),
    }

    return {
        "method_name": "Holt-Winters (dengan musiman)",
        "method_short": "Holt-Winters",
        "mae": best["mae"], "rmse": best["rmse"], "mape": best["mape"],
        "params_info": params_info,
        "df_eval": df_eval, "df_tabel": df_tabel, "df_forecast": df_forecast,
        "tabel_kandidat": tabel_kandidat,
        "X": X, "Ft_plot": list(Ft_full), "n": n,
        "tgl_forecast": tgl_forecast, "hasil_forecast": hasil_forecast,
    }


# ──────────────────────────────────────────────────────────────────────────
# 4. PIPELINE DATA (preprocessing, sama untuk kedua metode)
# ──────────────────────────────────────────────────────────────────────────

def load_and_clean(file_obj):
    df = pd.read_excel(file_obj)

    if "TGL_PS" not in df.columns:
        st.error(f"File tidak punya kolom 'TGL_PS'. Kolom ditemukan: {df.columns.tolist()}")
        st.stop()

    df["TGL_PS"] = pd.to_datetime(df["TGL_PS"], format="%Y%m%d")

    catatan = []
    catatan.append(f"Total baris mentah   : {len(df):,}")
    catatan.append(f"Rentang tanggal mentah: {df['TGL_PS'].min().date()} s/d {df['TGL_PS'].max().date()}")

    tanggal_terakhir = df["TGL_PS"].max()
    periode_terakhir = tanggal_terakhir.to_period("M")
    jumlah_hari_bulan = periode_terakhir.days_in_month

    if tanggal_terakhir.day < jumlah_hari_bulan:
        baris_sebelum = len(df)
        df = df[df["TGL_PS"].dt.to_period("M") != periode_terakhir].copy()
        baris_sesudah = len(df)
        catatan.append(
            f"🧹 Data cleaning: bulan {periode_terakhir.strftime('%B %Y')} tidak lengkap "
            f"(baru sampai tanggal {tanggal_terakhir.day} dari {jumlah_hari_bulan} hari) "
            f"→ DIBUANG ({baris_sebelum - baris_sesudah:,} baris)."
        )
    else:
        catatan.append("✅ Bulan terakhir sudah lengkap, tidak ada baris yang dibuang.")

    harian = df.groupby("TGL_PS").size().reset_index(name="jumlah").sort_values("TGL_PS")
    full_range = pd.date_range(harian["TGL_PS"].min(), harian["TGL_PS"].max(), freq="D")
    harian = harian.set_index("TGL_PS").reindex(full_range, fill_value=0)
    harian.index.name = "TGL_PS"

    catatan.append(f"Total hari (bersih)   : {len(harian):,} hari")
    return harian, catatan


def build_monthly(harian, n_forecast):
    bulanan = harian["jumlah"].resample("ME").sum()
    info = (
        f"Periode data: {bulanan.index[0].strftime('%B %Y')} s/d "
        f"{bulanan.index[-1].strftime('%B %Y')} ({len(bulanan)} bulan) | "
        f"Horizon ramalan: {n_forecast} bulan ke depan"
    )
    return bulanan, info


def split_data(bulanan, test_ratio):
    n_test = round(len(bulanan) * test_ratio)
    n_train = len(bulanan) - n_test
    if n_train < 4:
        st.error(f"Data training terlalu sedikit ({n_train} bulan). Kecilkan rasio test atau gunakan data lebih panjang.")
        st.stop()
    train = bulanan.iloc[:n_train]
    test = bulanan.iloc[n_train:]
    info = (
        f"Training: {len(train)} bulan ({train.index[0].strftime('%b %Y')}–{train.index[-1].strftime('%b %Y')}) | "
        f"Testing: {len(test)} bulan ({test.index[0].strftime('%b %Y')}–{test.index[-1].strftime('%b %Y')})"
    )
    return train, test, info


def uji_stasioneritas(bulanan):
    hasil_adf = adfuller(bulanan.values)
    stat, pval = hasil_adf[0], hasil_adf[1]
    if pval < 0.05:
        kesimpulan = "Data STASIONER (p < 0.05)."
    else:
        kesimpulan = "Data TIDAK STASIONER (p ≥ 0.05) → ada tren."
    return stat, pval, kesimpulan


def buat_plot(X, Ft_plot, n, n_train, tgl_forecast, hasil_forecast, method_short, mae, rmse, mape):
    fig, ax = plt.subplots(figsize=(12, 5.5))

    ax.plot(range(1, n + 1), X, color="steelblue", linewidth=2, marker="o", markersize=4, label="Data Aktual")

    idx_fit = [i + 1 for i in range(n) if Ft_plot[i] is not None]
    val_fit = [Ft_plot[i] for i in range(n) if Ft_plot[i] is not None]
    ax.plot(idx_fit, val_fit, color="orange", linewidth=1.3, linestyle="--",
            marker="s", markersize=3, label=f"Fitted ({method_short})")

    idx_forecast = range(n + 1, n + len(hasil_forecast) + 1)
    ax.plot(idx_forecast, hasil_forecast, color="red", linewidth=2, linestyle="--",
            marker="D", markersize=5, label=f"Peramalan {len(hasil_forecast)} Bulan")

    for idx, val in zip(idx_forecast, hasil_forecast):
        ax.annotate(str(val), xy=(idx, val), xytext=(0, 8), textcoords="offset points",
                    ha="center", fontsize=8, color="red", fontweight="bold")

    ax.axvline(x=n_train + 0.5, color="gray", linestyle=":", linewidth=1.3, label="Batas Train/Test")
    ax.axvline(x=n + 0.5, color="black", linestyle=":", linewidth=1.3, label="Batas Data/Peramalan")

    ax.set_title(
        f"Peramalan Jumlah Pemasangan — {method_short}\n"
        f"MAE={mae:.1f}  RMSE={rmse:.1f}  MAPE={mape:.2f}% ({kategori_mape(mape)})",
        fontsize=11, fontweight="bold"
    )
    ax.set_xlabel("Periode (Bulan ke-)")
    ax.set_ylabel("Jumlah Pemasangan")
    ax.legend(fontsize=8, loc="upper left")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    return fig


def export_excel(hasil, n_train, n_test):
    wb = Workbook()
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")

    def sh(cell, color="1F4E79"):
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = center
        cell.border = bdr

    def sd(cell, alt=False):
        if alt:
            cell.fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        cell.alignment = center
        cell.border = bdr

    ws0 = wb.active
    ws0.title = "Ringkasan"
    ringkasan = [("Metode", hasil["method_name"]), ("Data Training", f"{n_train} bulan"),
                 ("Data Testing", f"{n_test} bulan")]
    ringkasan += list(hasil["params_info"].items())
    ringkasan += [("MAE (test)", round(hasil["mae"], 3)), ("RMSE (test)", round(hasil["rmse"], 3)),
                  ("MAPE (test) %", round(hasil["mape"], 3)), ("Kategori", kategori_mape(hasil["mape"]))]
    for i, (k, v) in enumerate(ringkasan, 1):
        c1 = ws0.cell(row=i, column=1, value=k); c1.font = Font(bold=True); c1.border = bdr
        c2 = ws0.cell(row=i, column=2, value=v); c2.border = bdr
    ws0.column_dimensions["A"].width = 20
    ws0.column_dimensions["B"].width = 32

    ws2 = wb.create_sheet("Evaluasi Model")
    for col, h in enumerate(hasil["df_eval"].columns, 1):
        sh(ws2.cell(row=1, column=col, value=h))
    for i, row in hasil["df_eval"].iterrows():
        for col, v in enumerate(row, 1):
            sd(ws2.cell(row=i + 2, column=col, value=v), alt=(i % 2 == 0))
    for col, w in enumerate([16, 10, 10, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws3 = wb.create_sheet("Tabel Perhitungan")
    for col, h in enumerate(hasil["df_tabel"].columns, 1):
        sh(ws3.cell(row=1, column=col, value=h))
    for i, row in hasil["df_tabel"].iterrows():
        for col, v in enumerate(row, 1):
            sd(ws3.cell(row=i + 2, column=col, value=v), alt=(i % 2 == 0))
    for col, w in enumerate([5, 10, 12, 12, 10], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    ws4 = wb.create_sheet("Hasil Peramalan")
    for col, h in enumerate(hasil["df_forecast"].columns, 1):
        sh(ws4.cell(row=1, column=col, value=h))
    for i, row in hasil["df_forecast"].iterrows():
        for col, v in enumerate(row, 1):
            sd(ws4.cell(row=i + 2, column=col, value=v), alt=(i % 2 == 0))
    for col, w in enumerate([6, 18, 12], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    if "tabel_mape" in hasil:
        ws5 = wb.create_sheet("Optimasi Parameter (MAPE)")
        tabel_mape = hasil["tabel_mape"]
        ws5.cell(row=1, column=1, value="α \\ γ")
        sh(ws5.cell(row=1, column=1))
        gamma_vals = list(tabel_mape.columns)
        alpha_vals = list(tabel_mape.index)
        for j, g in enumerate(gamma_vals, 2):
            sh(ws5.cell(row=1, column=j, value=g))
        for i, a in enumerate(alpha_vals, 2):
            sh(ws5.cell(row=i, column=1, value=a))
            for j, g in enumerate(gamma_vals, 2):
                c = ws5.cell(row=i, column=j, value=tabel_mape.loc[a, g])
                sd(c, alt=(i % 2 == 0))
                if a == hasil["best_alpha"] and g == hasil["best_gamma"]:
                    c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    c.font = Font(bold=True, color="C00000")
        for col in range(1, 11):
            ws5.column_dimensions[get_column_letter(col)].width = 9

    if "tabel_kandidat" in hasil:
        ws5 = wb.create_sheet("Kandidat Holt-Winters")
        for col, h in enumerate(hasil["tabel_kandidat"].columns, 1):
            sh(ws5.cell(row=1, column=col, value=h))
        for i, row in hasil["tabel_kandidat"].iterrows():
            for col, v in enumerate(row, 1):
                sd(ws5.cell(row=i + 2, column=col, value=v), alt=(i % 2 == 0))
        for col, w in enumerate([16, 14, 10, 10, 10], 1):
            ws5.column_dimensions[get_column_letter(col)].width = w

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    with open(tmp.name, "rb") as f:
        return f.read()


# ──────────────────────────────────────────────────────────────────────────
# 5. ANTARMUKA STREAMLIT
# ──────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="Peramalan WiFi IndiBiz", layout="wide")

st.title("📡 Peramalan Jumlah Pemasangan WiFi IndiBiz")
st.markdown(
    "Upload file **Data Mentah (.xlsx)** dengan kolom `TGL_PS` (format tanggal `YYYYMMDD`), "
    "pilih metode & parameter di sidebar, lalu klik **Jalankan Analisis**."
)

with st.sidebar:
    st.header("⚙️ Pengaturan")
    uploaded_file = st.file_uploader("Upload Data Mentah (.xlsx)", type=["xlsx"])

    metode = st.selectbox(
        "Metode Peramalan",
        ["Holt's Double Exponential Smoothing (tanpa musiman)", "Holt-Winters (dengan musiman)"],
        help=(
            "Holt's DES: cuma level + tren, cocok kalau polanya naik/turun umum tanpa "
            "pengulangan bulanan yang jelas.\n\n"
            "Holt-Winters: level + tren + musiman, cocok kalau ada pola berulang tiap tahun "
            "di bulan yang sama — tapi butuh data minimal 2 tahun penuh untuk hasil stabil."
        ),
    )

    test_ratio = st.slider("Proporsi Data Testing", 0.10, 0.30, 0.20, 0.05)
    n_forecast = st.slider("Horizon Peramalan (bulan ke depan)", 1, 12, 6, 1)
    run_btn = st.button("🚀 Jalankan Analisis", type="primary", use_container_width=True)

if run_btn:
    if uploaded_file is None:
        st.error("Silakan upload file Data Mentah (.xlsx) terlebih dahulu.")
        st.stop()

    with st.spinner("Memuat & membersihkan data..."):
        harian, catatan_cleaning = load_and_clean(uploaded_file)

    with st.spinner("Menentukan periode peramalan..."):
        bulanan, info_periode = build_monthly(harian, n_forecast)

    with st.spinner("Membagi data train-test..."):
        train, test, info_split = split_data(bulanan, test_ratio)

    with st.spinner("Uji stasioneritas (ADF)..."):
        adf_stat, adf_pval, adf_kesimpulan = uji_stasioneritas(bulanan)

    if metode.startswith("Holt's"):
        with st.spinner("Optimasi parameter α, γ (81 kombinasi)..."):
            hasil = run_holt_des(bulanan, train, test, n_forecast)
    else:
        with st.spinner("Mencoba kombinasi musiman & damped trend (Holt-Winters)..."):
            hasil = run_holt_winters(bulanan, train, test, n_forecast)
        if hasil is None:
            st.error(
                f"Data training ({len(train)} bulan) belum cukup untuk Holt-Winters "
                f"(butuh minimal 24 bulan / 2 siklus musiman penuh). "
                f"Kecilkan rasio data testing, atau pakai Holt's DES."
            )
            st.stop()

    with st.spinner("Membuat visualisasi..."):
        fig = buat_plot(hasil["X"], hasil["Ft_plot"], hasil["n"], len(train),
                         hasil["tgl_forecast"], hasil["hasil_forecast"],
                         hasil["method_short"], hasil["mae"], hasil["rmse"], hasil["mape"])

    with st.spinner("Menyusun file Excel..."):
        excel_bytes = export_excel(hasil, len(train), len(test))

    st.success(f"Analisis selesai — Metode: {hasil['method_name']}")

    st.subheader("🧹 Preprocessing")
    for c in catatan_cleaning:
        st.markdown(f"- {c}")

    st.subheader("🗓️ Periode & Pembagian Data")
    st.markdown(f"- {info_periode}")
    st.markdown(f"- {info_split}")

    st.subheader("🔍 Uji Stasioneritas (ADF)")
    st.markdown(f"ADF statistic = `{adf_stat:.4f}` | p-value = `{adf_pval:.4f}` — {adf_kesimpulan}")

    st.subheader(f"🏆 Model Terbaik — {hasil['method_name']}")
    cols = st.columns(len(hasil["params_info"]) + 3)
    for col, (k, v) in zip(cols, hasil["params_info"].items()):
        col.metric(k, v)
    cols[-3].metric("MAE (test)", f"{hasil['mae']:.2f}")
    cols[-2].metric("RMSE (test)", f"{hasil['rmse']:.2f}")
    cols[-1].metric("MAPE (test)", f"{hasil['mape']:.2f}%")
    st.caption(f"Kategori akurasi: **{kategori_mape(hasil['mape'])}**")

    if "tabel_kandidat" in hasil:
        with st.expander("Lihat perbandingan 4 kombinasi Holt-Winters yang dicoba"):
            st.dataframe(hasil["tabel_kandidat"], use_container_width=True, hide_index=True)

    st.subheader(f"🔮 Ramalan {n_forecast} Bulan ke Depan")
    st.markdown(
        f"Total: **{sum(hasil['hasil_forecast']):,} pemasangan** | "
        f"Rata-rata: **{sum(hasil['hasil_forecast'])/n_forecast:.1f} pemasangan/bulan**"
    )

    st.pyplot(fig)

    tab1, tab2, tab3 = st.tabs(["📅 Hasil Peramalan", "✅ Evaluasi Model (Test)", "📋 Tabel Perhitungan Lengkap"])
    with tab1:
        st.dataframe(hasil["df_forecast"], use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(hasil["df_eval"], use_container_width=True, hide_index=True)
    with tab3:
        st.dataframe(hasil["df_tabel"], use_container_width=True, hide_index=True)

    if "tabel_mape" in hasil:
        with st.expander("Tabel Optimasi Parameter (MAPE untuk 81 kombinasi α × γ)"):
            st.dataframe(hasil["tabel_mape"].style.highlight_min(axis=None, color="#c6efce"),
                         use_container_width=True)

    st.download_button(
        label="⬇️ Download Hasil Lengkap (Excel)",
        data=excel_bytes,
        file_name=f"hasil_peramalan_{hasil['method_short'].replace(' ', '_').lower()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
else:
    st.info("👈 Upload data, pilih metode, lalu klik **Jalankan Analisis** di sidebar untuk memulai.")

st.markdown("---")
st.caption(
    "Kedua metode dievaluasi dengan pendekatan train-test split (out-of-sample) sebelum "
    "di-refit ke seluruh data untuk peramalan akhir. Bandingkan MAPE dari kedua metode "
    "untuk menentukan mana yang lebih cocok untuk pola data kamu."
)
