# -*- coding: utf-8 -*-
"""
📡 Aplikasi Peramalan Jumlah Pemasangan WiFi IndiBiz
Metode: Holt's Double Exponential Smoothing
Deploy dengan Gradio.

Alur: Pengumpulan Data -> Preprocessing (Data Cleaning) -> Penentuan Periode ->
      Pembagian Data (Train-Test) -> Pemilihan Metode (ADF) -> Membangun Model ->
      Optimasi Parameter (grid search alpha/gamma) -> Evaluasi Model (MAE, RMSE, MAPE) ->
      Pemilihan Model Terbaik -> Peramalan ke Depan
"""

import io
import tempfile

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from statsmodels.tsa.stattools import adfuller
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import gradio as gr

# ──────────────────────────────────────────────────────────────────────────
# 1. FUNGSI-FUNGSI INTI (model & metrik)
# ──────────────────────────────────────────────────────────────────────────

def holt_des_fit(data, alpha, gamma):
    """Fitting Holt's Double Exponential Smoothing pada sebuah deret data."""
    X = list(data)
    n = len(X)
    St = [0.0] * n
    bt = [0.0] * n
    Ft = [None] * n

    St[0] = X[0]
    bt[0] = ((X[1] - X[0]) + (X[3] - X[2])) / 2

    for t in range(1, n):
        St[t] = alpha * X[t] + (1 - alpha) * (St[t - 1] + bt[t - 1])
        bt[t] = gamma * (St[t] - St[t - 1]) + (1 - gamma) * bt[t - 1]
        Ft[t] = St[t - 1] + bt[t - 1]

    return St, bt, Ft


def forecast_m_langkah(St_akhir, bt_akhir, m_langkah):
    return [St_akhir + bt_akhir * m for m in range(1, m_langkah + 1)]


def hitung_metrik(aktual, ramalan):
    aktual = np.array(aktual, dtype=float)
    ramalan = np.array(ramalan, dtype=float)
    error = aktual - ramalan

    mae = np.mean(np.abs(error))
    rmse = np.sqrt(np.mean(error ** 2))

    mask = aktual != 0
    mape = np.mean(np.abs(error[mask] / aktual[mask])) * 100 if mask.sum() > 0 else np.nan

    return mae, rmse, mape


def kategori_mape(mape):
    if mape <= 10:
        return "Sangat Baik"
    elif mape <= 20:
        return "Baik"
    elif mape <= 50:
        return "Cukup Baik"
    else:
        return "Buruk"


# ──────────────────────────────────────────────────────────────────────────
# 2. PIPELINE DATA (dipanggil oleh UI Gradio)
# ──────────────────────────────────────────────────────────────────────────

def load_and_clean(file_path):
    """Tahap 1-2: Pengumpulan Data & Preprocessing (Data Cleaning)."""
    df = pd.read_excel(file_path)

    kolom_wajib = {"TGL_PS"}
    if not kolom_wajib.issubset(set(df.columns)):
        raise gr.Error(
            f"File tidak punya kolom yang diharapkan ({kolom_wajib}). "
            f"Kolom ditemukan: {df.columns.tolist()}"
        )

    df["TGL_PS"] = pd.to_datetime(df["TGL_PS"], format="%Y%m%d")

    catatan = []
    catatan.append(f"Total baris mentah   : {len(df):,}")
    catatan.append(
        f"Rentang tanggal mentah: {df['TGL_PS'].min().date()} s/d {df['TGL_PS'].max().date()}"
    )

    # Data cleaning: buang bulan terakhir jika belum lengkap sebulan penuh
    tanggal_terakhir = df["TGL_PS"].max()
    periode_terakhir = tanggal_terakhir.to_period("M")
    jumlah_hari_bulan = periode_terakhir.days_in_month

    if tanggal_terakhir.day < jumlah_hari_bulan:
        baris_sebelum = len(df)
        df = df[df["TGL_PS"].dt.to_period("M") != periode_terakhir].copy()
        baris_sesudah = len(df)
        catatan.append(
            f"🧹 Data cleaning: bulan {periode_terakhir.strftime('%B %Y')} tidak lengkap "
            f"(baru sampai tanggal {tanggal_terakhir.day} dari {jumlah_hari_bulan} hari) "
            f"→ DIBUANG ({baris_sebelum - baris_sesudah:,} baris)."
        )
    else:
        catatan.append("✅ Bulan terakhir sudah lengkap, tidak ada baris yang dibuang.")

    # Agregasi harian + isi tanggal kosong dengan 0
    harian = df.groupby("TGL_PS").size().reset_index(name="jumlah").sort_values("TGL_PS")
    full_range = pd.date_range(harian["TGL_PS"].min(), harian["TGL_PS"].max(), freq="D")
    harian = harian.set_index("TGL_PS").reindex(full_range, fill_value=0)
    harian.index.name = "TGL_PS"

    catatan.append(f"Total hari (bersih)   : {len(harian):,} hari")

    return harian, catatan


def build_monthly(harian, n_forecast):
    """Tahap 3: Menentukan Periode/Waktu Peramalan (agregasi bulanan)."""
    bulanan = harian["jumlah"].resample("ME").sum()
    info = (
        f"Periode data   : {bulanan.index[0].strftime('%B %Y')} s/d "
        f"{bulanan.index[-1].strftime('%B %Y')} ({len(bulanan)} bulan)\n"
        f"Horizon ramalan: {n_forecast} bulan ke depan"
    )
    return bulanan, info


def split_data(bulanan, test_ratio):
    """Tahap 4: Pembagian Data (Train-Test Split, berurutan sesuai waktu)."""
    n_test = round(len(bulanan) * test_ratio)
    n_train = len(bulanan) - n_test
    if n_train < 4:
        raise gr.Error(
            f"Data training terlalu sedikit ({n_train} bulan). "
            "Kecilkan rasio test atau gunakan data lebih panjang."
        )
    train = bulanan.iloc[:n_train]
    test = bulanan.iloc[n_train:]
    info = (
        f"Data Training : {len(train)} bulan ({train.index[0].strftime('%b %Y')} – "
        f"{train.index[-1].strftime('%b %Y')})\n"
        f"Data Testing  : {len(test)} bulan ({test.index[0].strftime('%b %Y')} – "
        f"{test.index[-1].strftime('%b %Y')})"
    )
    return train, test, info


def uji_stasioneritas(bulanan):
    """Tahap 5: Pemilihan Metode Peramalan (Uji ADF)."""
    hasil_adf = adfuller(bulanan.values)
    stat, pval = hasil_adf[0], hasil_adf[1]
    if pval < 0.05:
        kesimpulan = "Data STASIONER (p < 0.05). Holt's DES tetap relevan bila tren terlihat pada grafik."
    else:
        kesimpulan = "Data TIDAK STASIONER (p ≥ 0.05) → ada tren → Holt's DES sesuai digunakan."
    info = f"ADF statistic = {stat:.4f} | p-value = {pval:.4f}\n{kesimpulan}"
    return info


def optimasi_parameter(train, test):
    """Tahap 7: Optimasi Parameter (grid search 81 kombinasi α, γ pada test set)."""
    alpha_vals = [round(x * 0.1, 1) for x in range(1, 10)]
    gamma_vals = [round(x * 0.1, 1) for x in range(1, 10)]

    hasil_metrik = {}
    tabel_mape = pd.DataFrame(index=alpha_vals, columns=gamma_vals, dtype=float)

    for a in alpha_vals:
        for g in gamma_vals:
            St_tr, bt_tr, _ = holt_des_fit(train.values, a, g)
            ramalan_test = forecast_m_langkah(St_tr[-1], bt_tr[-1], len(test))
            mae, rmse, mape = hitung_metrik(test.values, ramalan_test)
            hasil_metrik[(a, g)] = (mae, rmse, mape)
            tabel_mape.loc[a, g] = round(mape, 2)

    best_key = min(hasil_metrik, key=lambda k: hasil_metrik[k][2])
    best_alpha, best_gamma = best_key
    best_mae, best_rmse, best_mape = hasil_metrik[best_key]

    return best_alpha, best_gamma, best_mae, best_rmse, best_mape, tabel_mape, hasil_metrik


def tabel_evaluasi(train, test, best_alpha, best_gamma):
    """Tahap 8: Evaluasi Model — detail per-bulan pada data test."""
    St_tr, bt_tr, _ = holt_des_fit(train.values, best_alpha, best_gamma)
    ramalan_test = forecast_m_langkah(St_tr[-1], bt_tr[-1], len(test))

    df_eval = pd.DataFrame({
        "Periode": [t.strftime("%B %Y") for t in test.index],
        "Aktual": test.values,
        "Ramalan": [round(v, 1) for v in ramalan_test],
    })
    df_eval["Error"] = (df_eval["Aktual"] - df_eval["Ramalan"]).round(1)
    df_eval["APE (%)"] = (
        df_eval["Error"].abs() / df_eval["Aktual"].replace(0, np.nan) * 100
    ).round(2)
    return df_eval


def refit_dan_ramal(bulanan, best_alpha, best_gamma, n_forecast):
    """Tahap 10-11: Refit ke seluruh data + peramalan ke depan."""
    St, bt, Ft = holt_des_fit(bulanan.values, best_alpha, best_gamma)
    X = bulanan.values.tolist()
    periode = bulanan.index.tolist()
    n = len(X)

    rows = []
    for t in range(n):
        ft = round(Ft[t], 2) if Ft[t] is not None else None
        pet = None
        if t > 0 and X[t] != 0 and Ft[t] is not None:
            pet = round(abs((X[t] - Ft[t]) / X[t]) * 100, 2)
        rows.append({
            "t": t + 1,
            "Periode": periode[t].strftime("%b %Y"),
            "Xt (Aktual)": X[t],
            "St (Level)": round(St[t], 2),
            "bt (Tren)": round(bt[t], 2),
            "Ft (Forecast)": ft,
            "PEt (%)": pet,
        })
    df_tabel = pd.DataFrame(rows)

    St_last, bt_last = St[-1], bt[-1]
    tgl_forecast = pd.date_range(
        start=bulanan.index[-1] + pd.DateOffset(months=1), periods=n_forecast, freq="ME"
    )
    hasil_forecast = [max(0, round(St_last + bt_last * m)) for m in range(1, n_forecast + 1)]

    df_forecast = pd.DataFrame({
        "m": range(1, n_forecast + 1),
        "Periode": [t.strftime("%B %Y") for t in tgl_forecast],
        "Prediksi": hasil_forecast,
    })

    return df_tabel, df_forecast, St, bt, Ft, X, n, tgl_forecast, hasil_forecast


def buat_plot(bulanan, n_train, X, Ft, n, tgl_forecast, hasil_forecast,
              best_alpha, best_gamma, best_mae, best_rmse, best_mape):
    fig, ax = plt.subplots(figsize=(12, 5.5))

    idx_aktual = range(1, n + 1)
    ax.plot(idx_aktual, X, color="steelblue", linewidth=2, marker="o", markersize=4,
            label="Data Aktual")

    Ft_plot = [Ft[t] for t in range(1, n)]
    ax.plot(range(2, n + 1), Ft_plot, color="orange", linewidth=1.3, linestyle="--",
            marker="s", markersize=3, label=f"Fitted (α={best_alpha}, γ={best_gamma})")

    idx_forecast = range(n + 1, n + len(hasil_forecast) + 1)
    ax.plot(idx_forecast, hasil_forecast, color="red", linewidth=2, linestyle="--",
            marker="D", markersize=5, label=f"Peramalan {len(hasil_forecast)} Bulan")

    for idx, val in zip(idx_forecast, hasil_forecast):
        ax.annotate(str(val), xy=(idx, val), xytext=(0, 8), textcoords="offset points",
                    ha="center", fontsize=8, color="red", fontweight="bold")

    ax.axvline(x=n_train + 0.5, color="gray", linestyle=":", linewidth=1.3, label="Batas Train/Test")
    ax.axvline(x=n + 0.5, color="black", linestyle=":", linewidth=1.3, label="Batas Data/Peramalan")

    ax.set_title(
        f"Peramalan Jumlah Pemasangan — Holt's DES\n"
        f"α={best_alpha}  γ={best_gamma}  |  MAE={best_mae:.1f}  RMSE={best_rmse:.1f}  "
        f"MAPE={best_mape:.2f}% ({kategori_mape(best_mape)})",
        fontsize=11, fontweight="bold"
    )
    ax.set_xlabel("Periode (Bulan ke-)")
    ax.set_ylabel("Jumlah Pemasangan")
    ax.legend(fontsize=8, loc="upper left")
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    return fig


def export_excel(df_tabel, df_eval, df_forecast, tabel_mape,
                  best_alpha, best_gamma, best_mae, best_rmse, best_mape,
                  n_train, n_test):
    wb = Workbook()
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")

    def sh(cell, color="1F4E79"):
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = center
        cell.border = bdr

    def sd(cell, alt=False):
        if alt:
            cell.fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        cell.alignment = center
        cell.border = bdr

    # Sheet 1: Ringkasan
    ws0 = wb.active
    ws0.title = "Ringkasan"
    ringkasan = [
        ("Metode", "Holt's Double Exponential Smoothing"),
        ("Data Training", f"{n_train} bulan"),
        ("Data Testing", f"{n_test} bulan"),
        ("Alpha terbaik", best_alpha),
        ("Gamma terbaik", best_gamma),
        ("MAE (test)", round(best_mae, 3)),
        ("RMSE (test)", round(best_rmse, 3)),
        ("MAPE (test) %", round(best_mape, 3)),
        ("Kategori", kategori_mape(best_mape)),
    ]
    for i, (k, v) in enumerate(ringkasan, 1):
        c1 = ws0.cell(row=i, column=1, value=k)
        c1.font = Font(bold=True)
        c1.border = bdr
        c2 = ws0.cell(row=i, column=2, value=v)
        c2.border = bdr
    ws0.column_dimensions["A"].width = 20
    ws0.column_dimensions["B"].width = 30

    # Sheet 2: Optimasi Parameter (MAPE)
    ws1 = wb.create_sheet("Optimasi Parameter (MAPE)")
    ws1.cell(row=1, column=1, value="α \\ γ")
    sh(ws1.cell(row=1, column=1))
    gamma_vals = list(tabel_mape.columns)
    alpha_vals = list(tabel_mape.index)
    for j, g in enumerate(gamma_vals, 2):
        sh(ws1.cell(row=1, column=j, value=g))
    for i, a in enumerate(alpha_vals, 2):
        sh(ws1.cell(row=i, column=1, value=a))
        for j, g in enumerate(gamma_vals, 2):
            c = ws1.cell(row=i, column=j, value=tabel_mape.loc[a, g])
            sd(c, alt=(i % 2 == 0))
            if a == best_alpha and g == best_gamma:
                c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                c.font = Font(bold=True, color="C00000")
    for col in range(1, 11):
        ws1.column_dimensions[get_column_letter(col)].width = 9

    # Sheet 3: Evaluasi Model (test set)
    ws2 = wb.create_sheet("Evaluasi Model")
    for col, h in enumerate(df_eval.columns, 1):
        sh(ws2.cell(row=1, column=col, value=h))
    for i, row in df_eval.iterrows():
        for col, v in enumerate(row, 1):
            c = ws2.cell(row=i + 2, column=col, value=v)
            sd(c, alt=(i % 2 == 0))
    for col, w in enumerate([16, 10, 10, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Sheet 4: Tabel Perhitungan DES (seluruh data)
    ws3 = wb.create_sheet("Tabel Perhitungan DES")
    for col, h in enumerate(df_tabel.columns, 1):
        sh(ws3.cell(row=1, column=col, value=h))
    for i, row in df_tabel.iterrows():
        for col, v in enumerate(row, 1):
            c = ws3.cell(row=i + 2, column=col, value=v)
            sd(c, alt=(i % 2 == 0))
    for col, w in enumerate([5, 10, 12, 10, 10, 12, 10], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # Sheet 5: Hasil Peramalan
    ws4 = wb.create_sheet("Hasil Peramalan")
    for col, h in enumerate(df_forecast.columns, 1):
        sh(ws4.cell(row=1, column=col, value=h))
    for i, row in df_forecast.iterrows():
        for col, v in enumerate(row, 1):
            c = ws4.cell(row=i + 2, column=col, value=v)
            sd(c, alt=(i % 2 == 0))
    for col, w in enumerate([6, 18, 12], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name


# ──────────────────────────────────────────────────────────────────────────
# 3. FUNGSI UTAMA YANG DIPANGGIL TOMBOL "JALANKAN ANALISIS"
# ──────────────────────────────────────────────────────────────────────────

def jalankan_analisis(file_obj, test_ratio, n_forecast, progress=gr.Progress()):
    if file_obj is None:
        raise gr.Error("Silakan upload file Data Mentah (.xlsx) terlebih dahulu.")

    progress(0.05, desc="Memuat & membersihkan data...")
    harian, catatan_cleaning = load_and_clean(file_obj)

    progress(0.15, desc="Menentukan periode peramalan...")
    bulanan, info_periode = build_monthly(harian, n_forecast)

    progress(0.25, desc="Membagi data train-test...")
    train, test, info_split = split_data(bulanan, test_ratio)

    progress(0.35, desc="Uji stasioneritas (ADF)...")
    info_adf = uji_stasioneritas(bulanan)

    progress(0.45, desc="Optimasi parameter α, γ (81 kombinasi)...")
    (best_alpha, best_gamma, best_mae, best_rmse, best_mape,
     tabel_mape, _) = optimasi_parameter(train, test)

    progress(0.65, desc="Evaluasi model pada data test...")
    df_eval = tabel_evaluasi(train, test, best_alpha, best_gamma)

    progress(0.75, desc="Refit model & meramalkan ke depan...")
    (df_tabel, df_forecast, St, bt, Ft, X, n,
     tgl_forecast, hasil_forecast) = refit_dan_ramal(bulanan, best_alpha, best_gamma, n_forecast)

    progress(0.85, desc="Membuat visualisasi...")
    fig = buat_plot(bulanan, len(train), X, Ft, n, tgl_forecast, hasil_forecast,
                     best_alpha, best_gamma, best_mae, best_rmse, best_mape)

    progress(0.95, desc="Menyusun file Excel...")
    excel_path = export_excel(df_tabel, df_eval, df_forecast, tabel_mape,
                               best_alpha, best_gamma, best_mae, best_rmse, best_mape,
                               len(train), len(test))

    ringkasan_md = f"""
### 🧹 Preprocessing
{chr(10).join('- ' + c for c in catatan_cleaning)}

### 🗓️ Periode & Pembagian Data
- {info_periode}
- {info_split.replace(chr(10), chr(10) + '- ')}

### 🔍 Uji Stasioneritas (ADF)
{info_adf}

### 🏆 Model Terbaik (hasil optimasi 81 kombinasi α, γ)
| Parameter | Nilai |
|---|---|
| α (alpha) | **{best_alpha}** |
| γ (gamma) | **{best_gamma}** |
| MAE (test) | **{best_mae:.2f}** |
| RMSE (test) | **{best_rmse:.2f}** |
| MAPE (test) | **{best_mape:.2f}%** |
| Kategori akurasi | **{kategori_mape(best_mape)}** |

### 🔮 Ramalan {n_forecast} Bulan ke Depan
Total: **{sum(hasil_forecast):,} pemasangan** | Rata-rata: **{sum(hasil_forecast)/n_forecast:.1f} pemasangan/bulan**
"""

    return (
        ringkasan_md,
        fig,
        df_forecast,
        df_eval,
        tabel_mape.reset_index().rename(columns={"index": "α \\ γ"}),
        excel_path,
    )


# ──────────────────────────────────────────────────────────────────────────
# 4. ANTARMUKA GRADIO
# ──────────────────────────────────────────────────────────────────────────

with gr.Blocks(title="Peramalan WiFi IndiBiz — Holt's DES") as demo:
    gr.Markdown(
        """
        # 📡 Peramalan Jumlah Pemasangan WiFi IndiBiz
        ### Metode: Holt's Double Exponential Smoothing
        Upload file **Data Mentah (.xlsx)** dengan kolom `TGL_PS` (format tanggal `YYYYMMDD`),
        lalu klik **Jalankan Analisis**. Aplikasi ini otomatis membersihkan data (membuang bulan
        yang belum lengkap), membagi data train-test, mengoptimasi parameter α dan γ, mengevaluasi
        model (MAE, RMSE, MAPE), lalu meramalkan beberapa bulan ke depan.
        """
    )

    with gr.Row():
        with gr.Column(scale=1):
            file_input = gr.File(label="Upload Data Mentah (.xlsx)", file_types=[".xlsx"])
            test_ratio = gr.Slider(0.1, 0.3, value=0.2, step=0.05,
                                    label="Proporsi Data Testing", info="Default 20% data terakhir")
            n_forecast = gr.Slider(1, 12, value=6, step=1,
                                    label="Horizon Peramalan (bulan ke depan)")
            run_btn = gr.Button("🚀 Jalankan Analisis", variant="primary")

        with gr.Column(scale=2):
            summary_out = gr.Markdown(label="Ringkasan Hasil")

    plot_out = gr.Plot(label="Visualisasi Data Aktual, Fitted, dan Peramalan")

    with gr.Row():
        forecast_table = gr.Dataframe(label="Tabel Hasil Peramalan", interactive=False)
        eval_table = gr.Dataframe(label="Tabel Evaluasi Model (Data Test)", interactive=False)

    with gr.Accordion("Tabel Optimasi Parameter (MAPE untuk 81 kombinasi α × γ)", open=False):
        mape_table = gr.Dataframe(label="MAPE (%) per kombinasi α, γ", interactive=False)

    excel_out = gr.File(label="⬇️ Download Hasil Lengkap (Excel)")

    run_btn.click(
        fn=jalankan_analisis,
        inputs=[file_input, test_ratio, n_forecast],
        outputs=[summary_out, plot_out, forecast_table, eval_table, mape_table, excel_out],
    )

    gr.Markdown(
        "---\n*Aplikasi ini menggunakan pendekatan train-test split (out-of-sample) untuk "
        "memilih parameter α dan γ terbaik, sebelum model di-refit ke seluruh data untuk "
        "peramalan akhir.*"
    )

if __name__ == "__main__":
    demo.launch()
